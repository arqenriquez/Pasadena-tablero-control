# -*- coding: utf-8 -*-
"""Genera los Excel de metas del módulo PPC de Pasadena (plantilla + semana 09)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

BASE = os.path.join(os.path.dirname(__file__), "..", "data", "ppc")
HEADERS = ["Actividad", "Responsable", "Partida"]

ACCENT = "5A9BD4"   # azul cielo Pasadena
SOFT = "EAF3FB"

def estilizar(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

def escribir(ws, filas):
    thin = Side(style="thin", color="ECECEC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ri, fila in enumerate(filas, start=2):
        for ci, val in enumerate(fila, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci == 1))
            c.border = border
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FAFCFE")

# ---- Plantilla en blanco ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metas"
estilizar(ws)
escribir(ws, [
    ["Ejemplo: Colado de cimentación ciclópea", "", "Cimentación"],
    ["Ejemplo: Suministro y habilitado de PTRs", "", "Estructura techumbre"],
    ["", "", ""],
    ["", "", ""],
])
wb.save(os.path.join(BASE, "_plantilla-metas.xlsx"))

# ---- Semana 09 (primera semana en curso) ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Metas Semana 09"
estilizar(ws)
escribir(ws, [
    ["Continuación de muro de block hasta altura de cubierta", "", "Albañilería"],
    ["Continuación de colado de castillos", "", "Albañilería"],
    ["Block dala (cerramiento)", "", "Albañilería"],
    ["Colado de cimentación ciclópea", "", "Cimentación"],
    ["Suministro y habilitado de PTRs", "", "Estructura techumbre"],
])
wb.save(os.path.join(BASE, "metas", "semana-09.xlsx"))

print("Excel generados:")
print(" -", os.path.normpath(os.path.join(BASE, "_plantilla-metas.xlsx")))
print(" -", os.path.normpath(os.path.join(BASE, "metas", "semana-09.xlsx")))
